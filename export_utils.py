# -*- coding: utf-8 -*-
"""
export_utils.py
ส่งออกรายงานการเช็คชื่อเป็นไฟล์ Excel (.xlsx)
เงื่อนไข: นับว่า "เข้าร่วมกิจกรรม" เฉพาะคนที่มีทั้งสแกนเข้าและสแกนออกเท่านั้น
"""

import os
import tempfile
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import config
import database
import chart_utils


def _style_header(ws, row_idx, color):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def export_attendance_report(event_id, event_name):
    report = database.get_attendance_report(event_id)  # ทุกคนที่สแกนอย่างน้อย 1 ครั้ง (เข้าหรือออก)
    absent = database.get_absent_report(event_id)       # ไม่สแกนเลยทั้งเข้าและออก
    summary = database.get_attendance_summary(event_id)

    attended = [r for r in report if r["attended"]]
    not_attended_partial = [r for r in report if not r["attended"]]

    wb = Workbook()

    # ---------- ชีตสรุปภาพรวม (พร้อมกราฟ) ----------
    ws0 = wb.active
    ws0.title = "สรุปภาพรวม"
    ws0["A1"] = f"สรุปผลการเข้าร่วมกิจกรรม: {event_name}"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"] = f"วันที่ออกรายงาน: {datetime.date.today().isoformat()}"
    ws0["A3"] = "เงื่อนไข: นับเป็น \"เข้าร่วมกิจกรรม\" เฉพาะผู้ที่มีทั้งการสแกนเข้าและสแกนออกเท่านั้น"
    ws0["A3"].font = Font(italic=True, color="666666")

    summary_rows = [
        ("จำนวนนักเรียนทั้งหมด", summary["total"]),
        ("เข้าร่วมกิจกรรม (สแกนครบเข้า-ออก)", summary["attended"]),
        ("ไม่เข้าร่วมกิจกรรม (สแกนไม่ครบ/ไม่ได้สแกน)", summary["not_attended"]),
        ("  - สแกนเพียงฝั่งเดียว", summary["partial"]),
        ("  - ไม่ได้สแกนเลย", summary["absent"]),
    ]
    start_row = 5
    for i, (label, value) in enumerate(summary_rows):
        ws0.cell(row=start_row + i, column=1, value=label)
        ws0.cell(row=start_row + i, column=2, value=value)
    ws0.cell(row=start_row, column=1).font = Font(bold=True)
    ws0.cell(row=start_row + 1, column=1).font = Font(bold=True, color="2E7D32")
    ws0.cell(row=start_row + 2, column=1).font = Font(bold=True, color="C62828")

    # ฝังกราฟสรุป (วงกลม + แท่งตามห้อง + แท่งตามสาขา) ลงในชีทเดียวกัน
    try:
        fig = chart_utils.build_summary_figure(summary, event_name)
        tmp_path = os.path.join(tempfile.gettempdir(), f"chart_{event_id}.png")
        fig.savefig(tmp_path, dpi=150, bbox_inches="tight")
        img = XLImage(tmp_path)
        img.width = img.width * 0.5
        img.height = img.height * 0.5
        ws0.add_image(img, "A12")
    except Exception:
        pass  # ถ้าฝังกราฟไม่สำเร็จ (เช่นไม่มี matplotlib) ยังส่งออกตัวเลขสรุปได้ตามปกติ

    # ---------- ตารางสรุปแยกตามสาขา (ต่อจากกราฟ) ----------
    dept_start_row = 42
    ws0.cell(row=dept_start_row, column=1, value="สรุปแยกตามสาขา").font = Font(bold=True, size=12)
    header_row = dept_start_row + 1
    for col, text in enumerate(["สาขา", "เข้าร่วมกิจกรรม", "ไม่เข้าร่วมกิจกรรม", "รวม"], start=1):
        ws0.cell(row=header_row, column=col, value=text)
    _style_header(ws0, header_row, "2E5395")

    for i, (dept, counts) in enumerate(summary.get("by_department", {}).items(), start=1):
        r = header_row + i
        total_dept = counts["attended"] + counts["not_attended"]
        ws0.cell(row=r, column=1, value=dept)
        ws0.cell(row=r, column=2, value=counts["attended"])
        ws0.cell(row=r, column=3, value=counts["not_attended"])
        ws0.cell(row=r, column=4, value=total_dept)

    # ---------- ชีตเข้าร่วมกิจกรรม (ครบเข้า-ออก) ----------
    ws1 = wb.create_sheet("เข้าร่วมกิจกรรม")
    ws1.append([
        "ลำดับ", "รหัสนักเรียน", "ชื่อ-สกุล", "ห้อง", "สาขา",
        "เวลาสแกนเข้า", "เวลาสแกนออก", "ความมั่นใจเข้า (%)", "ความมั่นใจออก (%)",
    ])
    _style_header(ws1, 1, "2E7D32")
    for i, row in enumerate(attended, start=1):
        ws1.append([
            i, row["student_code"], row["full_name"], row["class_room"], row.get("department", ""),
            row["checkin_time"], row["check_out_time"],
            round((row["confidence"] or 0) * 100, 1),
            round((row["confidence_out"] or 0) * 100, 1),
        ])

    # ---------- ชีตไม่เข้าร่วมกิจกรรม (สแกนไม่ครบ + ไม่สแกนเลย) ----------
    ws2 = wb.create_sheet("ไม่เข้าร่วมกิจกรรม")
    ws2.append(["ลำดับ", "รหัสนักเรียน", "ชื่อ-สกุล", "ห้อง", "สาขา", "หมายเหตุ", "เวลาสแกนเข้า", "เวลาสแกนออก"])
    _style_header(ws2, 1, "C62828")

    row_num = 1
    for row in not_attended_partial:
        ws2.append([
            row_num, row["student_code"], row["full_name"], row["class_room"], row.get("department", ""),
            row["status_text"].replace("ไม่เข้าร่วมกิจกรรม ", "").strip("()"),
            row["checkin_time"] or "-", row["check_out_time"] or "-",
        ])
        row_num += 1
    for row in absent:
        ws2.append([
            row_num, row["student_code"], row["full_name"], row["class_room"], row.get("department", ""),
            "ไม่ได้สแกนเลย", "-", "-",
        ])
        row_num += 1

    for ws in (ws0, ws1, ws2):
        _autosize_columns(ws)
    ws0.column_dimensions["A"].width = 45  # ชีทสรุปมีข้อความยาว กำหนดกว้างพิเศษ

    safe_name = "".join(c for c in event_name if c not in '\\/:*?"<>|')
    filename = f"{safe_name}_{datetime.date.today().isoformat()}.xlsx"
    out_path = os.path.join(config.EXPORT_DIR, filename)
    wb.save(out_path)
    return out_path
