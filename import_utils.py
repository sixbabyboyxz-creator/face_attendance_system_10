# -*- coding: utf-8 -*-
"""
import_utils.py
นำเข้ารายชื่อนักเรียนจากไฟล์ Excel (.xlsx) หรือ CSV (.csv) เพื่อเตรียมข้อมูล
ก่อนลงทะเบียนใบหน้า (ไม่ต้องพิมพ์ทีละคนสำหรับนักเรียนจำนวนมาก)

รองรับ 3 รูปแบบไฟล์ (เรียงลำดับที่จะลองตรวจจับ):

  รูปแบบ A - ตารางแบนแบบเทมเพลตของระบบนี้:
    แถวแรกเป็นหัวตาราง เช่น "รหัสนักเรียน", "ชื่อ-สกุล", "ห้อง" (ไทย/อังกฤษก็ได้)
    แถวถัดไปเป็นข้อมูลตรงๆ

  รูปแบบ B - รายงานรายชื่อจากระบบ สอศ. (งานทะเบียน):
    มีแถวหัวเรื่อง + ข้อมูลกลุ่มเรียน (เช่น "ชั้นปี", "ปวส.1/1") แทรกอยู่ก่อน
    แล้วค่อยเจอแถวหัวตารางจริงที่มีคำว่า "ลำดับ" นำหน้า เช่น
      ลำดับ | รหัสประจำตัว | ชื่อ - สกุล | ประเภทผู้เรียน | สถานะนักเรียนนักศึกษา
    รองรับหลายกลุ่มเรียนซ้อนกันในชีทเดียว หรือหลายชีทในไฟล์เดียว
    ค่า "ชั้นปี" ที่เจอก่อนหน้าตารางจะถูกใช้เป็นค่า "ห้อง" ของนักเรียนทุกคนในตารางนั้น
    ถ้ามีคอลัมน์สถานะ จะข้ามแถวที่สถานะเป็น พ้นสภาพ/ลาออก/ตัดออก ฯลฯ อัตโนมัติ

  รูปแบบ C - ไม่มีหัวตารางเลย: สมมติคอลัมน์เรียงเป็น รหัส, ชื่อ, ห้อง ตามลำดับ
"""

import os
import csv
from openpyxl import load_workbook, Workbook

import database

_CODE_KEYWORDS = ["รหัส", "code", "id"]
_NAME_KEYWORDS = ["ชื่อ", "name"]
_CLASS_KEYWORDS = ["ห้อง", "ระดับ", "class", "room"]
_STATUS_KEYWORDS = ["สถานะ", "status"]
_TYPE_KEYWORDS = ["ประเภท", "type"]
_DEPT_KEYWORDS = ["สาขา", "แผนก", "department"]

_CLASS_LEVEL_MARKER = "ชั้นปี"
_GROUP_CODE_MARKER = "รหัสกลุ่มเรียน"
_GROUP_NAME_MARKER = "ชื่อกลุ่มเรียน"
_ADVISOR_MARKER = "ครูที่ปรึกษา"
_TABLE_HEADER_MARKER = "ลำดับ"

# สถานะที่ถือว่า "ไม่ได้เรียนแล้ว" ข้ามอัตโนมัติเวลานำเข้า (กันเอาคนที่พ้นสภาพเข้าระบบเช็คชื่อ)
_INACTIVE_STATUS_KEYWORDS = [
    "พ้นสภาพ", "ลาออก", "ตัดออก", "ให้พักการเรียน", "เสียชีวิต", "จำหน่าย", "หมดสภาพ",
]


def _cell_text(cell):
    return "" if cell is None else str(cell).strip()


def _match_header(cell_value, keywords):
    text = _cell_text(cell_value).lower()
    if not text:
        return False
    return any(k.lower() in text for k in keywords)


def _detect_flat_header(header_row):
    """รูปแบบ A: ตรวจแถวเดียวว่ามีหัวคอลัมน์ รหัส/ชื่อ/ห้อง/สาขา ครบไหม"""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if "code" not in mapping and _match_header(cell, _CODE_KEYWORDS):
            mapping["code"] = idx
        elif "name" not in mapping and _match_header(cell, _NAME_KEYWORDS):
            mapping["name"] = idx
        elif "class" not in mapping and _match_header(cell, _CLASS_KEYWORDS):
            mapping["class"] = idx
        elif "department" not in mapping and _match_header(cell, _DEPT_KEYWORDS):
            mapping["department"] = idx
    if "code" in mapping and "name" in mapping:
        return mapping
    return None


def _find_table_header_rows(rows):
    """รูปแบบ B: หาทุกแถวที่มีคำว่า 'ลำดับ' เป๊ะๆ (ใช้เป็นตัวบ่งชี้หัวตารางจริง กันสับสนกับ
    แถวข้อมูลกลุ่มเรียนอื่นๆ ที่อาจมีคำว่า 'รหัส'/'ชื่อ' ปนอยู่ เช่น 'รหัสกลุ่มเรียน')"""
    idxs = []
    for idx, row in enumerate(rows):
        for cell in row:
            if _cell_text(cell) == _TABLE_HEADER_MARKER:
                idxs.append(idx)
                break
    return idxs


def _find_preceding_marker_value(rows, before_idx, marker):
    """หาค่าที่อยู่ถัดจาก cell ที่ตรงกับ marker เป๊ะๆ (เช่น 'ชั้นปี', 'รหัสกลุ่มเรียน')
    โดยไล่หาจากแถวก่อนหน้า before_idx ถอยกลับไป (เอาค่าที่ใกล้ที่สุด/ล่าสุดก่อน)"""
    for row in reversed(rows[:before_idx]):
        for i, cell in enumerate(row):
            if _cell_text(cell) == marker and i + 1 < len(row):
                value = _cell_text(row[i + 1])
                if value:
                    return value
    return ""


def _is_blank_row(row):
    return all(_cell_text(c) == "" for c in row)


def _is_inactive_status(status_text):
    return any(k in status_text for k in _INACTIVE_STATUS_KEYWORDS)


def _read_rows_xlsx_all_sheets(path):
    wb = load_workbook(path, data_only=True)
    return [(ws.title, list(ws.iter_rows(values_only=True))) for ws in wb.worksheets]


def _read_rows_csv(path):
    for enc in ("utf-8-sig", "cp874", "utf-8"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                return [tuple(r) for r in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    raise ValueError("ไม่สามารถอ่านไฟล์ CSV ได้ (ปัญหาการเข้ารหัสตัวอักษร)")


def _process_flat_table(rows, mapping, start_row, sheet_label):
    """ประมวลผลตารางแบบ A/C (หัวตารางเดียวหรือไม่มีหัวตารางเลย)"""
    success, skipped, skipped_inactive, errors = 0, [], [], []
    code_idx, name_idx = mapping["code"], mapping["name"]
    class_idx = mapping.get("class")
    dept_idx = mapping.get("department")

    for i, row in enumerate(rows[start_row:], start=start_row + 1):
        if code_idx >= len(row) or name_idx >= len(row):
            continue
        code = _cell_text(row[code_idx])
        name = _cell_text(row[name_idx])
        klass = _cell_text(row[class_idx]) if (class_idx is not None and class_idx < len(row)) else ""
        dept = _cell_text(row[dept_idx]) if (dept_idx is not None and dept_idx < len(row)) else ""

        if not code or not name:
            continue

        if database.get_student_by_code(code):
            skipped.append(code)
            continue

        try:
            database.add_student(code, name, klass, department=dept)
            success += 1
        except Exception as e:
            errors.append(f"{sheet_label} แถวที่ {i} (รหัส {code}): {e}")

    return success, skipped, skipped_inactive, errors


def _process_report_table(rows, header_idx, next_boundary_idx, sheet_label):
    """ประมวลผลตารางแบบ B (รายงาน สอศ.) หนึ่งบล็อก ตั้งแต่ header_idx จนถึง next_boundary_idx"""
    header_row = rows[header_idx]
    code_idx = name_idx = status_idx = type_idx = None
    for idx, cell in enumerate(header_row):
        text = _cell_text(cell)
        if code_idx is None and _match_header(text, _CODE_KEYWORDS):
            code_idx = idx
        elif name_idx is None and _match_header(text, _NAME_KEYWORDS):
            name_idx = idx
        elif status_idx is None and _match_header(text, _STATUS_KEYWORDS):
            status_idx = idx
        elif type_idx is None and _match_header(text, _TYPE_KEYWORDS):
            type_idx = idx

    success, skipped, skipped_inactive, errors = 0, [], [], []

    if code_idx is None or name_idx is None:
        errors.append(f"{sheet_label}: พบแถวหัวตาราง ('ลำดับ') แต่หาคอลัมน์รหัส/ชื่อไม่เจอ")
        return success, skipped, skipped_inactive, errors

    class_room = _find_preceding_marker_value(rows, header_idx, _CLASS_LEVEL_MARKER)
    group_code = _find_preceding_marker_value(rows, header_idx, _GROUP_CODE_MARKER)
    group_name = _find_preceding_marker_value(rows, header_idx, _GROUP_NAME_MARKER)
    advisor = _find_preceding_marker_value(rows, header_idx, _ADVISOR_MARKER)

    # บันทึกข้อมูล "กลุ่มเรียน" แยกไว้ครั้งเดียวต่อบล็อก (ไม่ใช่ข้อมูลของนักเรียนรายคน)
    if group_code or group_name or advisor or class_room:
        database.get_or_create_class_group(
            group_code=group_code, group_name=group_name, level=class_room, advisor_teacher=advisor
        )

    for i in range(header_idx + 1, next_boundary_idx):
        row = rows[i]
        if _is_blank_row(row):
            break
        if code_idx >= len(row) or name_idx >= len(row):
            continue

        code = _cell_text(row[code_idx])
        name = _cell_text(row[name_idx])
        status = _cell_text(row[status_idx]) if status_idx is not None and status_idx < len(row) else ""
        student_type = _cell_text(row[type_idx]) if type_idx is not None and type_idx < len(row) else ""

        if not code or not name:
            continue

        if status and _is_inactive_status(status):
            skipped_inactive.append(code)
            continue

        if database.get_student_by_code(code):
            skipped.append(code)
            continue

        try:
            database.add_student(
                code, name, class_room, student_type=student_type, status=status, department=group_name,
            )
            success += 1
        except Exception as e:
            errors.append(f"{sheet_label} แถวที่ {i + 1} (รหัส {code}): {e}")

    return success, skipped, skipped_inactive, errors


def _import_rows(rows, sheet_label):
    success, skipped, skipped_inactive, errors = 0, [], [], []
    if not rows:
        return success, skipped, skipped_inactive, errors

    # --- ลองรูปแบบ A ก่อน: แถวแรกเป็นหัวตารางแบบง่าย ---
    flat_mapping = _detect_flat_header(rows[0])
    if flat_mapping is not None:
        return _process_flat_table(rows, flat_mapping, start_row=1, sheet_label=sheet_label)

    # --- ลองรูปแบบ B: หาแถวหัวตารางที่มีคำว่า "ลำดับ" (รองรับหลายบล็อกในชีทเดียว) ---
    header_idxs = _find_table_header_rows(rows)
    if header_idxs:
        for block_i, header_idx in enumerate(header_idxs):
            next_boundary = header_idxs[block_i + 1] if block_i + 1 < len(header_idxs) else len(rows)
            s, sk, ski, err = _process_report_table(rows, header_idx, next_boundary, sheet_label)
            success += s
            skipped += sk
            skipped_inactive += ski
            errors += err
        return success, skipped, skipped_inactive, errors

    # --- รูปแบบ C: ไม่มีหัวตารางเลย สมมติเรียงเป็น รหัส, ชื่อ, ห้อง ---
    mapping = {"code": 0, "name": 1, "class": 2}
    return _process_flat_table(rows, mapping, start_row=0, sheet_label=sheet_label)


def import_students_from_file(path):
    """
    นำเข้ารายชื่อนักเรียนจากไฟล์ .xlsx หรือ .csv ลงฐานข้อมูล
    รองรับทั้งไฟล์ตารางแบนธรรมดา และรายงานรายชื่อจากระบบ สอศ. (หลายชีท/หลายกลุ่มเรียนได้)
    ข้ามรหัสที่มีอยู่แล้วในระบบ และข้ามนักเรียนที่สถานะพ้นสภาพ/ลาออก ฯลฯ อัตโนมัติ

    คืนค่า dict: {
        "success": int,
        "skipped": list[str],           # รหัสที่ข้าม เพราะมีอยู่แล้วในระบบ
        "skipped_inactive": list[str],   # รหัสที่ข้าม เพราะสถานะไม่ได้เรียนแล้ว
        "errors": list[str],
    }
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        sheets = _read_rows_xlsx_all_sheets(path)
    elif ext == ".csv":
        sheets = [("CSV", _read_rows_csv(path))]
    else:
        raise ValueError("รองรับเฉพาะไฟล์ .xlsx หรือ .csv เท่านั้น")

    total_success = 0
    total_skipped = []
    total_skipped_inactive = []
    total_errors = []

    any_rows = False
    for sheet_name, rows in sheets:
        if not rows:
            continue
        any_rows = True
        s, sk, ski, err = _import_rows(rows, sheet_label=f"[{sheet_name}]")
        total_success += s
        total_skipped += sk
        total_skipped_inactive += ski
        total_errors += err

    if not any_rows:
        total_errors.append("ไฟล์ไม่มีข้อมูล")

    return {
        "success": total_success,
        "skipped": total_skipped,
        "skipped_inactive": total_skipped_inactive,
        "errors": total_errors,
    }


def create_template(path):
    """สร้างไฟล์ Excel ตัวอย่างสำหรับกรอกรายชื่อนักเรียนก่อนนำเข้า (รูปแบบ A แบบง่าย)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "รายชื่อนักเรียน"
    ws.append(["รหัสนักเรียน", "ชื่อ-สกุล", "ห้อง", "สาขา"])
    ws.append(["12345", "สมชาย ใจดี", "ปวช.1/1", "เทคโนโลยีสารสนเทศ"])
    ws.append(["12346", "สมหญิง ตั้งใจ", "ปวช.1/1", "เทคโนโลยีสารสนเทศ"])
    for col, width in zip("ABCD", (15, 30, 15, 25)):
        ws.column_dimensions[col].width = width
    wb.save(path)
